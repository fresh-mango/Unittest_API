#!/usr/bin/env python
# -*- encoding: utf-8 -*-
import os
import re
import openpyxl
from utils.config import ParserConfig,base_path,testcase_excel_path
from utils.operate_json import Operate_Json

# Testcase_Path = os.path.join(ParserConfig().get_path(), 'data','Testcase.xlsx')


class Operate_Excel():

    def __init__(self,file_name=None):
        if file_name==None:
            self.file_name = testcase_excel_path
        else:
            self.file_name = file_name
        self.open_excel = openpyxl.load_workbook(self.file_name)
        self.data =self.get_sheet_data()


    #加载所有的sheet内容
    def get_sheet_data(self,index=None):
        sheet_name = self.open_excel.sheetnames
        if index == None:
            index = 0
        data = self.open_excel[sheet_name[index]]
        return data


    #获取行数
    def get_rows(self):
        row = self.data.max_row
        return row


    #获取某一行的内容
    def get_rows_value(self,row):
        row_list = []
        for i in self.data[row]:
            row_list.append(i.value)
        return row_list

    #获取某一列的内容
    def get_columns_value(self,key=None):
        columns_list = []
        if key==None:
            key='A'
            columns_list_data = self.data[key]
            for i in columns_list_data:
                columns_list.append(i.value)
            return columns_list

        else:
            pattern = "^[A-Z]+$"
            res = re.search(pattern, string=key)
            if res:
                columns_list_data = self.data[key]
                for i in columns_list_data:
                    columns_list.append(i.value)
                return columns_list

            else:
                return "请输入正确的列表字段信息!"




    # #获取某一列的内容
    # def get_columns_value(self,key=None):
    #     columns_list = []
    #     if key==None:
    #         key='A'
    #     columns_list_data = self.data[key]
    #     for i in columns_list_data:
    #         columns_list.append(i.value)
    #     return columns_list



    #获取某一个单元格内容
    def get_cell_value(self,row,cols):
        data = self.data.cell(row=row + 1,column=cols + 1).value
        return data





    #获取行号
    def get_rows_number(self,case_id):
        num=1
        cols_data = self.get_columns_value()
        for col_data in cols_data:
            if col_data == case_id:
                return num
            num = num + 1
        #return num


    #获取excel所有的数据
    def get_excel_data(self):
        data_list =[]
        for i in range(self.get_rows())[1:]:
            data_list.append(self.get_rows_value(i+1))
        return data_list


    #获取【是否执行】=“yes”所在行的全部数据
    def test_case_list_yes(self):
        case_list = []
        rows = self.get_rows()
        for row in range(rows)[1:]:
            data = self.get_rows_value(row+1)
            if data[4] in ('yes','init') :
                row_num = self.get_rows_number(data[0])
                data = self.get_rows_value(row_num)
                case_list.append(data)
        return case_list




    #写入数据
    def excel_write_data(self, row, cols, value):
        wb = self.open_excel
        wr = wb.active
        wr.cell(row, cols, value)
        wb.save(file_name)

    #获取预期结果数据
    def get_expect_result(self,row):
        pc  = ParserConfig()
        col = pc.get_expect_result()
        expect_result = self.get_cell_value(row,col)
        return expect_result

    #获取请求数据_关键字key
    def get_request_data(self,row):
        pc  = ParserConfig()
        col = pc.get_request_data()
        request_key_data = self.get_cell_value(row,col)
        return request_key_data


    #通过请求数据_关键字key拿到json数据
    def get_data_for_json(self,row):
        op_json = Operate_Json()
        request_data = op_json.get_data(self.get_request_data(row))
        return request_data



    #拆分单元格数据
    def split_data_one(self,row,col):
        # imooc_005>data:banner:id
        data = self.get_cell_value(row, col)
        rule_key = data.split(">")[0]
        rule_data = data.split(">")[1]
        return rule_key, rule_data


    #拆分单元格数据,方式二
    def split_data_two(self,data):
        # imooc_005>data:banner:id
        rule_key = data.split(">")[0]
        rule_data = data.split(">")[1]
        return rule_key, rule_data




if __name__ == "__main__":

    oe = Operate_Excel()
    data = oe.get_rows_number('USA-brand-0009')
    print(type(data),data)










